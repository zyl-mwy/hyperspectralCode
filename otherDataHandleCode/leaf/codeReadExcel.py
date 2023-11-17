from openpyxl import load_workbook

wb = load_workbook('10_15_黄芪 - Copy (2).xlsx')
sheets = wb.worksheets
print(sheets)
sheet1 = sheets[0]
# sheet1 = wb['Sheet']  # 也可以通过已知表名获取sheet
print(sheet1)
# cell_11 = sheet1.cell(1,1).value


row1 = []
print(sheet1[1])
for row in sheet1[1]:
    print(row)
    row1.append(row.value)
print(row1)

col1 = []
for col in sheet1['A']:
    col1.append(col.value)
print(col1)


rows = sheet1.rows
columns = sheet1.columns
print(rows)
print(columns)
print('')

for row in rows:
    # print(rows)
    print(row)
    row_val = [col.value for col in row]
    print(row_val)
print('')

for col in columns:
    print(col)
    col_val = [row.value for row in col]
    print(col_val)


import xlrd
import csv

def xlsx_to_csv():
    workbook = xlrd.open_workbook('1.xlsx')
    table = workbook.sheet_by_index(0)
    with codecs.open('1.csv', 'w', encoding='utf-8') as f:
        write = csv.writer(f)
        for row_num in range(table.nrows):
            row_value = table.row_values(row_num)
            write.writerow(row_value)

xlsx_to_csv()
